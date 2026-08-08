#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from common import ROOT, load

def main():
    out=ROOT/'dist'/'downloads'; out.mkdir(parents=True,exist_ok=True)
    concerts=load('concerts.json',[]); sets=load('setlists.json',[]); changes=load('changelog.json',[])
    wb=Workbook(); ws=wb.active; ws.title='Concerts'
    fields=['date','year','artist','eventType','city','venue','country','event','songsPlayed','setlistStatus','setLengthMin','setTime','curfew','tour','attendance','concertCapacity','soldOut','venueAddress','generalVenueCapacity','sourceUrl','setlistFmUrl']
    ws.append(fields)
    for c in concerts: ws.append([c.get(f) for f in fields])
    ss=wb.create_sheet('Setlists'); sf=['date','concertId','section','position','song','countsAsSong','status','sourceUrl']; ss.append(sf)
    for s in sets: ss.append([s.get(f) for f in sf])
    ch=wb.create_sheet('Changelog'); cf=['timestamp','date','concertId','type','field','old','new','source']; ch.append(cf)
    for x in changes[-5000:]: ch.append([x.get(f) for f in cf])
    for sh in wb.worksheets:
        sh.freeze_panes='A2'; sh.auto_filter.ref=sh.dimensions
        for cell in sh[1]: cell.fill=PatternFill('solid',fgColor='20252B'); cell.font=Font(color='FFFFFF',bold=True); cell.alignment=Alignment(wrap_text=True)
        for col in sh.columns:
            letter=col[0].column_letter; sh.column_dimensions[letter].width=min(42,max(12,len(str(col[0].value or ''))+3))
    wb.save(out/'the-cure-tours-latest.xlsx'); print('[excel] written')
if __name__=='__main__': main()
