#!/usr/bin/env python3
from __future__ import annotations
import csv, json, re
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
XLSX=ROOT/'source'/'cure_concerts_V3_complete.xlsx'
CSV=ROOT/'source'/'cure_setlists_V3_complete.csv'
DATA=ROOT/'data'; DATA.mkdir(exist_ok=True)

def slug(s):
    s=(s or '').strip().lower()
    s=re.sub(r'[^a-z0-9]+','-',s)
    return s.strip('-')

def normalize_key(v):
    return re.sub(r'[^a-z0-9]+',' ',str(v or '').lower()).strip()

def norm(v):
    if v is None: return None
    if isinstance(v,bool): return v
    s=str(v).strip()
    return s if s else None

def main():
    wb=load_workbook(XLSX, read_only=True, data_only=True)
    ws=wb['Concerts']
    rows=ws.iter_rows(values_only=True)
    headers=[str(x) for x in next(rows)]
    concerts=[]
    for i,row in enumerate(rows,1):
        d=dict(zip(headers,row))
        date=norm(d.get('Date'))
        if not date: continue
        base_id=f"cureguide:{date}:{slug(norm(d.get('City')) or '')}:{slug(norm(d.get('Venue')) or '')}"
        cid=base_id
        suffix=2
        existing_ids={x['id'] for x in concerts}
        while cid in existing_ids:
            cid=f'{base_id}:{suffix}'; suffix+=1
        c={
            'id': cid, 'date':date, 'year':int(d.get('Year') or str(date)[:4]),
            'artist':norm(d.get('Artist / Project')) or 'The Cure', 'eventType':norm(d.get('Event Type')) or 'Concert',
            'city':norm(d.get('City')), 'venue':norm(d.get('Venue')), 'country':norm(d.get('Country')),
            'event':norm(d.get('Event / Festival')), 'songsPlayed':d.get('Songs Played'),
            'setlistStatus':norm(d.get('Setlist Status')), 'setLengthMin':d.get('Set Length (min)'),
            'setTime':norm(d.get('Set Time')), 'curfew':norm(d.get('Curfew')), 'dayOfWeek':norm(d.get('Day of Week')),
            'tour':norm(d.get('Tour')), 'attendance':d.get('Attendance'), 'concertCapacity':d.get('Concert Capacity'),
            'soldOut':norm(d.get('Sold Out')), 'venueAddress':norm(d.get('Venue Address')),
            'generalVenueCapacity':d.get('General Venue Capacity'), 'sourceUrl':norm(d.get('Source URL')),
            'pageTitle':norm(d.get('Page Title')), 'scrapedAt':norm(d.get('Scraped At')),
            'sources': {'primary':'cure-concerts.de'}
        }
        concerts.append(c)
    keymap={}
    for c in concerts:
        key=(c['date'], normalize_key(c.get('city')), normalize_key(c.get('venue')))
        keymap.setdefault(key,[]).append(c['id'])
    with open(CSV,encoding='utf-8-sig',newline='') as f:
        r=csv.DictReader(f)
        setlists=[]
        for x in r:
            key=(x['Date'], normalize_key(x.get('City')), normalize_key(x.get('Venue')))
            ids=keymap.get(key) or [c['id'] for c in concerts if c['date']==x['Date']]
            cid=ids[0] if ids else f"cureguide:{x['Date']}"
            setlists.append({
                'concertId':cid, 'date':x['Date'], 'section':x['Section'],
                'position':int(x['Position']) if x['Position'].isdigit() else None, 'song':x['Song'],
                'countsAsSong':x['Counts as Song'].lower()=='yes', 'status':x['Setlist Status'], 'sourceUrl':x['Source URL']
            })
    (DATA/'concerts.json').write_text(json.dumps(concerts,ensure_ascii=False,indent=2),encoding='utf-8')
    (DATA/'setlists.json').write_text(json.dumps(setlists,ensure_ascii=False,indent=2),encoding='utf-8')
    if not (DATA/'changelog.json').exists(): (DATA/'changelog.json').write_text('[]\n',encoding='utf-8')
    if not (DATA/'state.json').exists(): (DATA/'state.json').write_text(json.dumps({'lastSync':None,'version':'4.0.0'},indent=2),encoding='utf-8')
    print(f'Seeded {len(concerts)} concerts, {len(setlists)} setlist rows')
if __name__=='__main__': main()
